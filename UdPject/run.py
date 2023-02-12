'''
Project   name  :Recording data
code written by :MBONEA MJEMA
Email           :mjema86@gmail.com
2017
'''
import sys
import serial
from serial.serialutil import SerialException
import random
import openpyxl
import datetime
import time
from openpyxl.chart import (
ScatterChart,
Reference,
Series,
)
print('ok')
from serial import Serial

#usb port
Usb_Port="/dev/ttyUSB0"

#use this if the one above failed
# Usb_Port="/dev/ttyUSB1"   

#test serial port 
def Test_Serial_Port () :
    try:
        xbee=Serial(Usb_Port, 9600)
        xbee.close()
    except SerialException :
        print("Error Check the usb_port of the xbee or change the 'Usb_port' variale")
        sys.exit() 

  
    
try:
    RecordBook  = openpyxl.load_workbook("Records.xlsx")#load the workbook
#get the respective sheet by its name
    WATER=RecordBook.get_sheet_by_name("WATER")
    GAS=RecordBook.get_sheet_by_name("GAS")
except:
    #incase the file does not exist
    #this will create the file and the sheets
    RecordBook   = openpyxl.Workbook()
    WATER=RecordBook.create_sheet("WATER",index=0)
    GAS=RecordBook.create_sheet("GAS",index=1)
    for name in RecordBook.get_sheet_names():
        sheet=RecordBook.get_sheet_by_name(name)
        sheet[ 'A1']="DATE"
        sheet[ 'B1']="TIME"
        sheet[ 'C1']="VALUE"

#this contains the xbee_name and the sheet that has to be written        
XBEE_NAMES={1:("WATER\n",WATER),2:("GAS\n",GAS)}    

#get data gets data from the xbee with the particular name
def getData(Name_Of_RXbee):
    xbee=Serial(Usb_Port, 9600) # opening the port connected to the xbeee coordinator
    xbee.write(Name_Of_RXbee.encode()) #calling the router xbees
    data=xbee.readline()#receive data from the xbee which was called
    print("received")
    xbee.flush()
    xbee.close()
    print(float(data.decode().strip()))
    return float(data.decode().strip())#decode the data

#draws the chart from the data recorded
def DrawChart(sheet,title):
    chart = ScatterChart()
    chart.title = title
    chart.style = 10    
    chart.x_axis.title = 'Time'
    chart.y_axis.title = 'Mass in kg'
    xvalues = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
    yvalues = Reference(sheet, min_col=3, min_row=1, max_row=sheet.max_row)
    series = Series(yvalues, xvalues,title_from_data=True)
    chart.series.append(series)
    sheet.add_chart(chart, "D2")


#writes data in the file
def Write_sheet(data,Name_of_sheet):#it takes in data and the name of the sheet
    row=Name_of_sheet.max_row+1 
    Time=datetime.datetime.now().strftime("%H:%M")
    Date=datetime.date.today()
    Name_of_sheet.cell(row=row,column=1).value=Date#writes data
    Name_of_sheet.cell(row=row,column=2).value=Time
    Name_of_sheet.cell(row=row,column=3).value=data
    RecordBook.save(filename="Records.xlsx")

#the main method
def main():

    Test_Serial_Port ()
    print("ready...")    
    print("sending names....")

    #runs the program
    while True:
        for name in XBEE_NAMES:
            
            Xbee_name,sheet=XBEE_NAMES[name]
            data=getData(Xbee_name)#calls the name of the xbee and get the data
            print("recording data....")
            Write_sheet(data,sheet)#writes it in the sheet

        #draw the chart
        DrawChart(WATER,"WATER CONSUMPTION")
        DrawChart(GAS,"GAS COMSUMPTION")
        #save the file
        RecordBook.save(filename="Records.xlsx")
        #close the file
        RecordBook.close()
        print("done!!")
        time.sleep(30)#delay the program for 30 minutes
if __name__ == '__main__':
    main()
